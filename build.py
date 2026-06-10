"""Build the OTA Vietnam geo experiment analysis workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Style helpers
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', start_color='1F4E78')
INPUT_FONT = Font(name='Arial', color='0000FF', size=11)
FORMULA_FONT = Font(name='Arial', color='000000', size=11)
LINK_FONT = Font(name='Arial', color='008000', size=11)
BOLD = Font(name='Arial', bold=True, size=11)
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='1F4E78')
SECTION_FONT = Font(name='Arial', bold=True, size=12, color='1F4E78')
NORMAL = Font(name='Arial', size=11)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
HIGHLIGHT = PatternFill('solid', start_color='FFF2CC')
GREEN_BG = PatternFill('solid', start_color='E2EFDA')
RED_BG = PatternFill('solid', start_color='FCE4D6')
GREY_BG = PatternFill('solid', start_color='F2F2F2')

thin = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# TAB 1: README
# ============================================================
ws = wb.active
ws.title = 'README'

ws['A1'] = 'OTA Vietnam — Geo Experiment Analysis Workbook'
ws['A1'].font = TITLE_FONT

ws['A3'] = 'THE SCENARIO'
ws['A3'].font = SECTION_FONT

ws['A4'] = ("A fictional OTA currently spends $10,000 per week on Google Display ads in each of several Vietnamese cities. "
            "The Vietnam Country Manager wants to know whether increasing display spend by 50% (to $15,000/week) "
            "in those cities will generate enough incremental bookings to justify the extra spend.")
ws['A4'].alignment = WRAP
ws.merge_cells('A4:F4')
ws.row_dimensions[4].height = 60

ws['A6'] = 'THE EXPERIMENT DESIGN'
ws['A6'].font = SECTION_FONT

design_text = [
    "• 8 mid-sized Vietnamese cities, matched into 4 pairs based on size and historical booking patterns",
    "• Within each pair, one city is randomly assigned to TREATMENT (gets 50% spend increase), other to CONTROL (no change)",
    "• Pre-period: 4 weeks of observation BEFORE the intervention (to verify parallel trends)",
    "• Test period: 6 weeks of observation AFTER the intervention",
    "• Primary metric: incremental bookings per week",
    "• Analysis method: Difference-in-Differences (DiD)",
]
for i, line in enumerate(design_text, start=7):
    ws.cell(row=i, column=1, value=line).alignment = WRAP
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

ws['A14'] = 'WORKBOOK STRUCTURE'
ws['A14'].font = SECTION_FONT

tabs = [
    ('01_Setup', 'City list, group assignments (treatment/control), pair structure, and global parameters (spend, booking value, commission rate)'),
    ('02_Bookings', 'Raw weekly bookings data for all 8 cities across all 10 weeks (4 pre + 6 test)'),
    ('03_Spend', 'Weekly ad spend for all 8 cities across all 10 weeks'),
    ('04_Analysis', 'The Difference-in-Differences calculation. This is where you compute the answer.'),
    ('05_Sanity_Check', 'Verifies parallel trends in the pre-period (the foundation of DiD validity)'),
]
ws.cell(row=15, column=1, value='Tab').font = BOLD
ws.cell(row=15, column=2, value='What it contains').font = BOLD
for i, (tab, desc) in enumerate(tabs, start=16):
    ws.cell(row=i, column=1, value=tab).font = BOLD
    ws.cell(row=i, column=2, value=desc).alignment = WRAP
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)

ws['A22'] = 'SUGGESTED EXERCISES (interview-style)'
ws['A22'].font = SECTION_FONT

exercises = [
    "1. Check parallel trends — do treatment and control move together during the 4-week pre-period? (Tab 05)",
    "2. Compute average weekly bookings for treatment and control, separately for pre-period and test period. (Tab 04)",
    "3. Calculate the change (test avg − pre avg) for each group.",
    "4. Compute the Difference-in-Differences: (treatment change) − (control change) = incremental bookings/week.",
    "5. Compute the incremental cost per booking: incremental spend ÷ incremental bookings.",
    "6. Compute the incremental ROAS using commission revenue (booking value × commission rate).",
    "7. Compare your incremental ROAS to a typical platform-attributed ROAS — what's the takeaway?",
    "8. What caveats would you flag when presenting this to the Country Manager?",
]
for i, line in enumerate(exercises, start=23):
    ws.cell(row=i, column=1, value=line).alignment = WRAP
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

ws['A32'] = 'COLOR CONVENTIONS'
ws['A32'].font = SECTION_FONT
ws['A33'] = 'Blue text = hardcoded input (data, assumptions)'
ws['A33'].font = INPUT_FONT
ws['A34'] = 'Black text = formula / calculation'
ws['A34'].font = FORMULA_FONT
ws['A35'] = 'Green text = link to another tab'
ws['A35'].font = LINK_FONT
ws['A36'] = 'Yellow background = the cells you should compute / fill in'
ws['A36'].fill = HIGHLIGHT

set_col_widths(ws, [22, 25, 25, 25, 25, 25])

# ============================================================
# TAB 2: 01_Setup
# ============================================================
ws = wb.create_sheet('01_Setup')

ws['A1'] = 'Experiment Setup'
ws['A1'].font = TITLE_FONT

ws['A3'] = 'City Assignments'
ws['A3'].font = SECTION_FONT

setup_headers = ['Pair', 'City', 'Group', 'Region', 'Pre-period weekly baseline (bookings)']
for j, h in enumerate(setup_headers, 1):
    c = ws.cell(row=4, column=j, value=h)
    header(c)

cities = [
    (1, 'Da Nang', 'Treatment', 'Central Coast', 3200),
    (1, 'Hai Phong', 'Control', 'Northern Coast', 3400),
    (2, 'Can Tho', 'Treatment', 'Mekong Delta', 3000),
    (2, 'Nha Trang', 'Control', 'Southern Coast', 3100),
    (3, 'Hue', 'Treatment', 'Central', 2800),
    (3, 'Vung Tau', 'Control', 'Southern Coast', 3000),
    (4, 'Quy Nhon', 'Treatment', 'Central Coast', 2800),
    (4, 'Buon Ma Thuot', 'Control', 'Central Highlands', 2900),
]
for i, row in enumerate(cities, start=5):
    for j, val in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=val)
        c.font = INPUT_FONT
        c.border = BORDER
        if j == 3:
            c.fill = GREEN_BG if val == 'Treatment' else RED_BG

# Parameters
ws['A15'] = 'Global Parameters'
ws['A15'].font = SECTION_FONT

params = [
    ('Control weekly spend per city ($)', 10000, 'The baseline display spend in each city'),
    ('Treatment weekly spend per city ($)', 15000, 'The +50% spend in treatment cities during test period'),
    ('Number of treatment cities', 4, ''),
    ('Number of control cities', 4, ''),
    ('Pre-period weeks', 4, 'Weeks of observation before intervention'),
    ('Test period weeks', 6, 'Weeks of observation after intervention'),
    ('Average booking value ($)', 120, 'Average hotel night price in Vietnam'),
    ('Commission rate (%)', 0.15, 'OTA commission as % of booking value'),
    ('Revenue per booking ($)', '=B23*B24', 'AOV × commission rate'),
]
ws.cell(row=16, column=1, value='Parameter').font = BOLD
ws.cell(row=16, column=2, value='Value').font = BOLD
ws.cell(row=16, column=3, value='Notes').font = BOLD
for c in ws[16]:
    c.fill = GREY_BG
    c.border = BORDER
for i, (name, val, note) in enumerate(params, start=17):
    ws.cell(row=i, column=1, value=name).font = NORMAL
    cv = ws.cell(row=i, column=2, value=val)
    if isinstance(val, str) and val.startswith('='):
        cv.font = FORMULA_FONT
    else:
        cv.font = INPUT_FONT
    if 'rate' in name.lower():
        cv.number_format = '0.0%'
    elif 'value' in name.lower() or 'spend' in name.lower() or 'Revenue' in name:
        cv.number_format = '$#,##0'
    ws.cell(row=i, column=3, value=note).font = NORMAL
    for j in range(1, 4):
        ws.cell(row=i, column=j).border = BORDER

set_col_widths(ws, [38, 14, 50, 22, 30])

# ============================================================
# TAB 3: 02_Bookings
# ============================================================
ws = wb.create_sheet('02_Bookings')

ws['A1'] = 'Weekly Bookings Data'
ws['A1'].font = TITLE_FONT
ws['A2'] = 'Weeks -4 to -1 = pre-period (no change). Weeks 1 to 6 = test period (treatment got +50% spend).'
ws['A2'].font = NORMAL

# Headers
booking_headers = ['City', 'Group', 'Week', 'Period', 'Bookings']
for j, h in enumerate(booking_headers, 1):
    c = ws.cell(row=4, column=j, value=h)
    header(c)

# Generate realistic data
# Design: each city has a baseline. Market lift factor (same for all). Treatment gets extra ad lift during test period.
# Idiosyncratic noise per (city, week) for realism but small.

import random
random.seed(7)

# Weekly market factor (applies to everyone — captures season/economy)
market_factor = {
    -4: 1.000, -3: 1.008, -2: 0.992, -1: 1.017,
    1:  1.045, 2:  1.060, 3:  1.070, 4:  1.085, 5:  1.095, 6:  1.115
}
# Treatment-only ad lift during test period (additive on top of market)
ad_lift = {
    -4: 0.000, -3: 0.000, -2: 0.000, -1: 0.000,
    1:  0.045, 2:  0.050, 3:  0.052, 4:  0.058, 5:  0.062, 6:  0.068
}
# Small noise per city-week
def noise():
    return random.uniform(-0.015, 0.015)

city_info = {
    'Da Nang': ('Treatment', 3200),
    'Hai Phong': ('Control', 3400),
    'Can Tho': ('Treatment', 3000),
    'Nha Trang': ('Control', 3100),
    'Hue': ('Treatment', 2800),
    'Vung Tau': ('Control', 3000),
    'Quy Nhon': ('Treatment', 2800),
    'Buon Ma Thuot': ('Control', 2900),
}
week_order = [-4, -3, -2, -1, 1, 2, 3, 4, 5, 6]

row = 5
for city, (group, base) in city_info.items():
    for w in week_order:
        period = 'Pre' if w < 0 else 'Test'
        lift = market_factor[w] + (ad_lift[w] if group == 'Treatment' else 0) + noise()
        bookings = round(base * lift)
        ws.cell(row=row, column=1, value=city).font = INPUT_FONT
        ws.cell(row=row, column=2, value=group).font = INPUT_FONT
        ws.cell(row=row, column=3, value=w).font = INPUT_FONT
        ws.cell(row=row, column=4, value=period).font = INPUT_FONT
        ws.cell(row=row, column=5, value=bookings).font = INPUT_FONT
        for j in range(1, 6):
            ws.cell(row=row, column=j).border = BORDER
            if j == 2:
                ws.cell(row=row, column=j).fill = GREEN_BG if group == 'Treatment' else RED_BG
            if j == 4:
                ws.cell(row=row, column=j).fill = GREY_BG if period == 'Pre' else HIGHLIGHT
        row += 1

set_col_widths(ws, [16, 13, 8, 10, 12])

# ============================================================
# TAB 4: 03_Spend
# ============================================================
ws = wb.create_sheet('03_Spend')

ws['A1'] = 'Weekly Ad Spend Data'
ws['A1'].font = TITLE_FONT
ws['A2'] = 'All cities spend $10,000/week during pre-period. Treatment cities switch to $15,000/week in test period; control cities stay at $10,000.'
ws['A2'].font = NORMAL

spend_headers = ['City', 'Group', 'Week', 'Period', 'Spend ($)']
for j, h in enumerate(spend_headers, 1):
    c = ws.cell(row=4, column=j, value=h)
    header(c)

row = 5
for city, (group, base) in city_info.items():
    for w in week_order:
        period = 'Pre' if w < 0 else 'Test'
        if period == 'Pre':
            spend = 10000
        else:
            spend = 15000 if group == 'Treatment' else 10000
        ws.cell(row=row, column=1, value=city).font = INPUT_FONT
        ws.cell(row=row, column=2, value=group).font = INPUT_FONT
        ws.cell(row=row, column=3, value=w).font = INPUT_FONT
        ws.cell(row=row, column=4, value=period).font = INPUT_FONT
        sc = ws.cell(row=row, column=5, value=spend)
        sc.font = INPUT_FONT
        sc.number_format = '$#,##0'
        for j in range(1, 6):
            ws.cell(row=row, column=j).border = BORDER
            if j == 2:
                ws.cell(row=row, column=j).fill = GREEN_BG if group == 'Treatment' else RED_BG
        row += 1

set_col_widths(ws, [16, 13, 8, 10, 12])

# ============================================================
# TAB 5: 04_Analysis
# ============================================================
ws = wb.create_sheet('04_Analysis')

ws['A1'] = 'Difference-in-Differences Analysis'
ws['A1'].font = TITLE_FONT
ws['A2'] = 'Walk through the DiD calculation step by step. Yellow cells are computed via formulas — inspect them to learn the logic.'
ws['A2'].font = NORMAL

# Step 1: Group totals per week
ws['A4'] = 'Step 1 — Total bookings per week, by group'
ws['A4'].font = SECTION_FONT

step1_headers = ['Week', 'Period', 'Treatment total bookings', 'Control total bookings']
for j, h in enumerate(step1_headers, 1):
    c = ws.cell(row=5, column=j, value=h)
    header(c)

# Each row is a week; use SUMIFS against the Bookings tab
for i, w in enumerate(week_order, start=6):
    period = 'Pre' if w < 0 else 'Test'
    ws.cell(row=i, column=1, value=w).font = INPUT_FONT
    ws.cell(row=i, column=2, value=period).font = INPUT_FONT
    # Treatment total
    tc = ws.cell(row=i, column=3, value=f'=SUMIFS(\'02_Bookings\'!E:E,\'02_Bookings\'!C:C,A{i},\'02_Bookings\'!B:B,"Treatment")')
    tc.font = LINK_FONT
    tc.number_format = '#,##0'
    tc.fill = HIGHLIGHT
    # Control total
    cc = ws.cell(row=i, column=4, value=f'=SUMIFS(\'02_Bookings\'!E:E,\'02_Bookings\'!C:C,A{i},\'02_Bookings\'!B:B,"Control")')
    cc.font = LINK_FONT
    cc.number_format = '#,##0'
    cc.fill = HIGHLIGHT
    for j in range(1, 5):
        ws.cell(row=i, column=j).border = BORDER

# Step 2: Averages by period
ws['A17'] = 'Step 2 — Average weekly bookings, by group and period'
ws['A17'].font = SECTION_FONT

step2_headers = ['', 'Treatment avg', 'Control avg']
for j, h in enumerate(step2_headers, 1):
    c = ws.cell(row=18, column=j, value=h)
    header(c)

ws.cell(row=19, column=1, value='Pre-period average (weeks -4 to -1)').font = BOLD
ws.cell(row=19, column=1).border = BORDER
t_pre = ws.cell(row=19, column=2, value='=AVERAGEIFS(C6:C15,B6:B15,"Pre")')
t_pre.font = FORMULA_FONT
t_pre.number_format = '#,##0'
t_pre.fill = HIGHLIGHT
t_pre.border = BORDER
c_pre = ws.cell(row=19, column=3, value='=AVERAGEIFS(D6:D15,B6:B15,"Pre")')
c_pre.font = FORMULA_FONT
c_pre.number_format = '#,##0'
c_pre.fill = HIGHLIGHT
c_pre.border = BORDER

ws.cell(row=20, column=1, value='Test period average (weeks 1 to 6)').font = BOLD
ws.cell(row=20, column=1).border = BORDER
t_test = ws.cell(row=20, column=2, value='=AVERAGEIFS(C6:C15,B6:B15,"Test")')
t_test.font = FORMULA_FONT
t_test.number_format = '#,##0'
t_test.fill = HIGHLIGHT
t_test.border = BORDER
c_test = ws.cell(row=20, column=3, value='=AVERAGEIFS(D6:D15,B6:B15,"Test")')
c_test.font = FORMULA_FONT
c_test.number_format = '#,##0'
c_test.fill = HIGHLIGHT
c_test.border = BORDER

# Step 3: Change for each group
ws['A22'] = 'Step 3 — Change in average bookings (test avg − pre avg)'
ws['A22'].font = SECTION_FONT

step3_headers = ['', 'Treatment change', 'Control change']
for j, h in enumerate(step3_headers, 1):
    c = ws.cell(row=23, column=j, value=h)
    header(c)

ws.cell(row=24, column=1, value='Change in weekly bookings').font = BOLD
ws.cell(row=24, column=1).border = BORDER
tch = ws.cell(row=24, column=2, value='=B20-B19')
tch.font = FORMULA_FONT
tch.number_format = '#,##0;(#,##0);-'
tch.fill = HIGHLIGHT
tch.border = BORDER
cch = ws.cell(row=24, column=3, value='=C20-C19')
cch.font = FORMULA_FONT
cch.number_format = '#,##0;(#,##0);-'
cch.fill = HIGHLIGHT
cch.border = BORDER

# Step 4: DiD
ws['A26'] = 'Step 4 — Difference-in-Differences = incremental bookings per week'
ws['A26'].font = SECTION_FONT
ws['A27'] = 'DiD = (Treatment change) − (Control change). This isolates the effect of the ad spend increase from market-wide forces.'
ws['A27'].font = NORMAL
ws['A27'].alignment = WRAP
ws.merge_cells('A27:D27')
ws.row_dimensions[27].height = 30

ws.cell(row=28, column=1, value='Incremental bookings per week (DiD)').font = BOLD
ws.cell(row=28, column=1).border = BORDER
did = ws.cell(row=28, column=2, value='=B24-C24')
did.font = FORMULA_FONT
did.number_format = '#,##0;(#,##0);-'
did.fill = HIGHLIGHT
did.border = BORDER

# Step 5: Money
ws['A30'] = 'Step 5 — Translate to money (incremental CPA and ROAS)'
ws['A30'].font = SECTION_FONT

money_rows = [
    ('Incremental spend per week ($)',
     "='01_Setup'!B18*'01_Setup'!B19 - '01_Setup'!B17*'01_Setup'!B19",
     '$#,##0', 'Extra spend in treatment = (treatment $/city − control $/city) × # treatment cities'),
    ('Incremental cost per booking ($)',
     '=B31/B28',
     '$#,##0.00', 'Incremental spend ÷ incremental bookings'),
    ('Revenue per booking ($)',
     "='01_Setup'!B25",
     '$#,##0.00', 'From Setup tab (AOV × commission)'),
    ('Incremental commission revenue per week ($)',
     '=B28*B33',
     '$#,##0', 'Incremental bookings × revenue per booking'),
    ('Incremental ROAS',
     '=B34/B31',
     '0.00"x"', 'Incremental revenue ÷ incremental spend. >1 means profitable on commission.'),
]

ws.cell(row=30, column=2).value = ''  # placeholder so headers align
for i, (name, formula, fmt, note) in enumerate(money_rows, start=31):
    ws.cell(row=i, column=1, value=name).font = BOLD
    c = ws.cell(row=i, column=2, value=formula)
    if formula.startswith("='01_Setup'"):
        c.font = LINK_FONT
    else:
        c.font = FORMULA_FONT
    c.number_format = fmt
    c.fill = HIGHLIGHT
    ws.cell(row=i, column=4, value=note).font = NORMAL
    ws.cell(row=i, column=4).alignment = WRAP
    for j in [1, 2, 4]:
        ws.cell(row=i, column=j).border = BORDER

# Step 6: Decision framing
ws['A37'] = 'Step 6 — What to recommend'
ws['A37'].font = SECTION_FONT
ws['A38'] = ('If incremental ROAS > 1.0x: the extra spend at least covers itself on commission. '
             'If it clears your hurdle rate (often 1.5–2.0x to account for non-commission costs), recommend scaling. '
             'Then consider a follow-up test at a larger increase (+100%) to find where incremental ROAS drops below the threshold.')
ws['A38'].alignment = WRAP
ws.merge_cells('A38:D38')
ws.row_dimensions[38].height = 60

set_col_widths(ws, [42, 18, 18, 55])

# ============================================================
# TAB 6: 05_Sanity_Check
# ============================================================
ws = wb.create_sheet('05_Sanity_Check')

ws['A1'] = 'Parallel Trends Sanity Check'
ws['A1'].font = TITLE_FONT
ws['A2'] = ('DiD is only valid if treatment and control moved together during the pre-period. '
            'If their week-over-week changes diverge before the intervention, the control is not a good counterfactual.')
ws['A2'].font = NORMAL
ws['A2'].alignment = WRAP
ws.merge_cells('A2:E2')
ws.row_dimensions[2].height = 30

headers = ['Week', 'Treatment total', 'Control total', 'Treatment WoW % change', 'Control WoW % change']
for j, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=j, value=h)
    header(c)

# Only the pre-period weeks
pre_weeks = [-4, -3, -2, -1]
for i, w in enumerate(pre_weeks, start=5):
    ws.cell(row=i, column=1, value=w).font = INPUT_FONT
    tc = ws.cell(row=i, column=2, value=f"=SUMIFS('02_Bookings'!E:E,'02_Bookings'!C:C,A{i},'02_Bookings'!B:B,\"Treatment\")")
    tc.font = LINK_FONT
    tc.number_format = '#,##0'
    cc = ws.cell(row=i, column=3, value=f"=SUMIFS('02_Bookings'!E:E,'02_Bookings'!C:C,A{i},'02_Bookings'!B:B,\"Control\")")
    cc.font = LINK_FONT
    cc.number_format = '#,##0'
    if i > 5:
        twow = ws.cell(row=i, column=4, value=f'=B{i}/B{i-1}-1')
        twow.font = FORMULA_FONT
        twow.number_format = '0.0%;(0.0%);-'
        cwow = ws.cell(row=i, column=5, value=f'=C{i}/C{i-1}-1')
        cwow.font = FORMULA_FONT
        cwow.number_format = '0.0%;(0.0%);-'
    for j in range(1, 6):
        ws.cell(row=i, column=j).border = BORDER

ws['A11'] = 'Pre-period averages'
ws['A11'].font = SECTION_FONT
ws.cell(row=12, column=1, value='Treatment pre avg').font = BOLD
ws.cell(row=12, column=2, value='=AVERAGE(B5:B8)').font = FORMULA_FONT
ws.cell(row=12, column=2).number_format = '#,##0'
ws.cell(row=13, column=1, value='Control pre avg').font = BOLD
ws.cell(row=13, column=2, value='=AVERAGE(C5:C8)').font = FORMULA_FONT
ws.cell(row=13, column=2).number_format = '#,##0'
ws.cell(row=14, column=1, value='Avg WoW % — Treatment').font = BOLD
ws.cell(row=14, column=2, value='=AVERAGE(D6:D8)').font = FORMULA_FONT
ws.cell(row=14, column=2).number_format = '0.00%'
ws.cell(row=15, column=1, value='Avg WoW % — Control').font = BOLD
ws.cell(row=15, column=2, value='=AVERAGE(E6:E8)').font = FORMULA_FONT
ws.cell(row=15, column=2).number_format = '0.00%'
ws.cell(row=16, column=1, value='Difference (Treatment − Control)').font = BOLD
ws.cell(row=16, column=2, value='=B14-B15').font = FORMULA_FONT
ws.cell(row=16, column=2).number_format = '0.00%'
ws.cell(row=16, column=2).fill = HIGHLIGHT
for r in range(12, 17):
    for j in [1, 2]:
        ws.cell(row=r, column=j).border = BORDER

ws['A18'] = 'Interpretation guide'
ws['A18'].font = SECTION_FONT
ws['A19'] = ('If "Difference (Treatment − Control)" in average week-over-week % change is small (well under 1%), '
             'parallel trends holds and DiD is trustworthy. If it is large or persistent in one direction, the assumption is shaky '
             'and you should re-match cities or reconsider the test.')
ws['A19'].alignment = WRAP
ws.merge_cells('A19:E19')
ws.row_dimensions[19].height = 45

set_col_widths(ws, [32, 16, 16, 22, 22])

# Save
out = '/home/claude/geo_workbook/ota_geo_experiment.xlsx'
wb.save(out)
print(f'Saved to {out}')
